"""
Heatmap selector with live edge fitting
=========================================
Draw a rectangle on the heatmap — the collapsed profile and
Erf+Gaussian fit update live as you drag.

Press  N  to load a new file
Press  Q  to quit

Usage:  python heatmap_selector.py
"""

import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.widgets import RectangleSelector
from scipy.optimize import curve_fit
from scipy.special import erf
from openpyxl import load_workbook

BG = '#f8f8f6'
B, T, R, A = '#185FA5', '#0F6E56', '#A32D2D', '#BA7517'

# ── Load ──────────────────────────────────────────────────────────────────────

def load(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = np.array(
        [[v if isinstance(v, (int, float)) else 0 for v in row]
         for row in ws.iter_rows(values_only=True)], dtype=float)
    wb.close()
    print(f"  Loaded {data.shape[0]} rows x {data.shape[1]} cols"
          f"  min={data.min():.0f}  max={data.max():.0f}")
    return data

# ── Model ─────────────────────────────────────────────────────────────────────

def erf_gauss(x, A_amp, B_amp, C, d, f):
    z = (2 * np.sqrt(np.log(2)) / f) * (d - x)
    return (A_amp * (1 + erf(z))
            + B_amp * np.exp(-np.log(16) / f**2 * (d - x)**2)
            + C)

def fit_edge(x_nm, y):
    d_guess = float(x_nm[np.argmax(np.abs(np.gradient(y)))])
    amp     = (y.max() - y.min()) / 2
    step    = np.mean(y[-len(y)//4:]) - np.mean(y[:len(y)//4])
    A_sign  = -1 if step > 0 else 1
    p0      = [A_sign * abs(amp), abs(amp) * 0.3, y.min(), d_guess, 8.0]
    bounds  = ([-np.inf,       0, -np.inf, x_nm.min(), 0.5],
               [ np.inf,  np.inf,  np.inf, x_nm.max(),
                 (x_nm.max() - x_nm.min()) / 2])
    try:
        popt, _ = curve_fit(erf_gauss, x_nm, y, p0=p0,
                            bounds=bounds, maxfev=10000)
        return popt
    except Exception:
        return None

# ── Session runner ────────────────────────────────────────────────────────────

def run_session(path, nm_per_px, h_nm, cmap):
    data  = load(path)
    title = os.path.basename(path)

    fig = plt.figure(figsize=(16, 7), facecolor=BG)
    fig.canvas.manager.set_window_title(title)
    gs  = gridspec.GridSpec(2, 3, figure=fig,
                            left=0.05, right=0.97,
                            top=0.92, bottom=0.10,
                            hspace=0.45, wspace=0.35)

    ax_heat    = fig.add_subplot(gs[:, 0])
    ax_profile = fig.add_subplot(gs[0, 1])
    ax_fit     = fig.add_subplot(gs[0, 2])
    ax_info    = fig.add_subplot(gs[1, 1:])

    for ax in [ax_heat, ax_profile, ax_fit, ax_info]:
        ax.set_facecolor(BG)
    ax_info.axis('off')

    im = ax_heat.imshow(data, cmap=cmap, aspect='auto', origin='upper',
                        vmin=np.percentile(data, 2),
                        vmax=np.percentile(data, 98))
    plt.colorbar(im, ax=ax_heat, label='Count', shrink=0.85)
    ax_heat.set_xlabel('Column (px)', fontsize=9)
    ax_heat.set_ylabel('Row (px)', fontsize=9)
    ax_heat.set_title('Draw rectangle across edge\nN = new file   Q = quit',
                      fontsize=9)

    prof_line,  = ax_profile.plot([], [], color=B, lw=1.4)
    edge_vline   = ax_profile.axvline(-1e9, color=R, lw=1, ls='--', alpha=0.7)
    ax_profile.set_xlabel('Position (nm)', fontsize=9)
    ax_profile.set_ylabel('Mean intensity', fontsize=9)
    ax_profile.set_title('Collapsed profile', fontsize=9)
    ax_profile.grid(True, lw=0.3, alpha=0.4)

    scat_fit  = ax_fit.scatter([], [], s=10, color=B, alpha=0.6)
    fit_line, = ax_fit.plot([], [], color=R, lw=2)
    ax_fit.set_xlabel('Position (nm)', fontsize=9)
    ax_fit.set_ylabel('Mean intensity', fontsize=9)
    ax_fit.set_title('Erf+Gauss fit', fontsize=9)
    ax_fit.grid(True, lw=0.3, alpha=0.4)

    info_text = ax_info.text(0.02, 0.95, 'Select a region to see results.',
                              transform=ax_info.transAxes,
                              fontsize=9, va='top', fontfamily='monospace',
                              color='#222')

    def update(r0, r1, c0, c1):
        region  = data[r0:r1+1, c0:c1+1]
        profile = np.mean(region, axis=0)
        x_nm    = np.arange(c0, c1+1, dtype=float) * nm_per_px

        prof_line.set_data(x_nm, profile)
        ax_profile.set_xlim(x_nm[0], x_nm[-1])
        ax_profile.set_ylim(profile.min() * 0.95, profile.max() * 1.05)
        ax_profile.set_title(f'Rows {r0}-{r1}  cols {c0}-{c1}', fontsize=9)

        popt = fit_edge(x_nm, profile)

        if popt is not None:
            A_amp, B_amp, C, d_nm, f_nm = popt
            theta = 90 - np.degrees(np.arctan(f_nm / h_nm))

            edge_vline.set_xdata([d_nm, d_nm])

            x_fit = np.linspace(x_nm[0], x_nm[-1], 500)
            y_fit = erf_gauss(x_fit, *popt)
            scat_fit.set_offsets(np.column_stack([x_nm, profile]))
            fit_line.set_data(x_fit, y_fit)
            ax_fit.set_xlim(x_nm[0], x_nm[-1])
            ax_fit.set_ylim(min(profile.min(), y_fit.min()) * 0.95,
                            max(profile.max(), y_fit.max()) * 1.05)
            ax_fit.set_title(f'f = {f_nm:.1f} nm   theta = {theta:.2f} deg',
                             fontsize=9)

            meets = 'YES' if theta >= 89.4 else 'NO'
            info_text.set_text(
                f"Pixel size    = {nm_per_px} nm/px\n"
                f"Height h      = {h_nm} nm\n"
                f"Edge pos d    = {d_nm:.1f} nm\n"
                f"FWHM f        = {f_nm:.2f} nm\n"
                f"A (Erf amp)   = {A_amp:.3f}\n"
                f"B (Gauss amp) = {B_amp:.3f}\n"
                f"C (baseline)  = {C:.3f}\n"
                f"Sidewall th   = {theta:.2f} deg\n"
                f"Meets >=89.4  = {meets}"
            )
        else:
            ax_fit.set_title('Fit failed — widen selection', fontsize=9)
            info_text.set_text(
                'Fit failed.\n'
                'Make sure the rectangle spans\n'
                'from one material region to the other\n'
                'across the edge perpendicularly.'
            )

        fig.canvas.draw_idle()

    def onselect(eclick, erelease):
        x1, x2 = sorted([eclick.xdata, erelease.xdata])
        y1, y2 = sorted([eclick.ydata, erelease.ydata])
        c0 = max(0, int(x1)); c1 = min(data.shape[1]-1, int(x2))
        r0 = max(0, int(y1)); r1 = min(data.shape[0]-1, int(y2))
        if c1 > c0 and r1 > r0:
            update(r0, r1, c0, c1)

    # Two-click selection: click top-left corner, then bottom-right
    clicks = []
    rect_patch = [None]

    def on_click(event):
        if event.inaxes != ax_heat:
            return
        clicks.append((int(event.xdata), int(event.ydata)))
        ax_heat.plot(event.xdata, event.ydata, '+', color='cyan',
                     ms=12, mew=2)
        fig.canvas.draw_idle()

        if len(clicks) == 2:
            (c0, r0), (c1, r1) = clicks
            c0, c1 = sorted([c0, c1])
            r0, r1 = sorted([r0, r1])
            c0 = max(0, c0); c1 = min(data.shape[1]-1, c1)
            r0 = max(0, r0); r1 = min(data.shape[0]-1, r1)

            if rect_patch[0] is not None:
                rect_patch[0].remove()
            import matplotlib.patches as mp
            rect_patch[0] = mp.Rectangle(
                (c0, r0), c1-c0, r1-r0,
                lw=1.5, edgecolor='cyan', facecolor='cyan', alpha=0.2)
            ax_heat.add_patch(rect_patch[0])
            fig.canvas.draw_idle()

            if c1 > c0 and r1 > r0:
                update(r0, r1, c0, c1)
            clicks.clear()

    fig.canvas.mpl_connect('button_press_event', on_click)

    result = {'action': None}

    def on_key(event):
        if event.key in ('q', 'Q'):
            result['action'] = 'quit'
            plt.close(fig)
        elif event.key in ('n', 'N'):
            result['action'] = 'new'
            plt.close(fig)

    fig.canvas.mpl_connect('key_press_event', on_key)
    plt.show()
    return result['action']

# ── Entry point ────────────────────────────────────────────────────────────────

def get_params():
    print("\n" + "─"*40)
    path = input("Excel file path: ").strip()
    while not os.path.exists(path):
        print("File not found.")
        path = input("Excel file path: ").strip()
    nm_per_px = float(input("Pixel size (nm/px): ").strip())
    h_nm      = float(input("Feature height h (nm): ").strip())
    cmap      = input("Colourmap [hot]: ").strip() or 'hot'
    return path, nm_per_px, h_nm, cmap


if __name__ == '__main__':
    path, nm_per_px, h_nm, cmap = get_params()
    while True:
        action = run_session(path, nm_per_px, h_nm, cmap)
        if action == 'quit' or action is None:
            print("Goodbye.")
            break
        elif action == 'new':
            path, nm_per_px, h_nm, cmap = get_params()