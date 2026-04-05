"""
Electron detector edge analysis
================================
1. Open heatmap
2. Select row band
3. Collapse rows to mean profile
4. Fit Erf + Gaussian to the edge
5. Report sidewall angle

Usage:  python edge_analysis.py
"""

import os
import sys
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.widgets import SpanSelector
from scipy.optimize import curve_fit
from scipy.special import erf
from openpyxl import load_workbook

BG = '#f8f8f6'
B, T, R, A = '#185FA5', '#0F6E56', '#A32D2D', '#BA7517'

# ── 1. Load ───────────────────────────────────────────────────────────────────

def load(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    data = np.array(
        [[v if isinstance(v, (int, float)) else 0 for v in row]
         for row in ws.iter_rows(values_only=True)], dtype=float)
    wb.close()
    print(f"  Loaded {data.shape[0]} rows x {data.shape[1]} cols")
    return data

# ── 2. Select rows ────────────────────────────────────────────────────────────

sel = [None, None]

def pick_rows(data):
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.imshow(data, cmap='hot', aspect='auto', origin='upper',
              vmin=np.percentile(data, 2), vmax=np.percentile(data, 98))
    ax.set_xlabel('Column (px)'); ax.set_ylabel('Row (px)')
    ax.set_title('Drag vertically to select row band over the edge — close when done')
    patch = [None]

    def onselect(y0, y1):
        r0, r1 = int(round(y0)), int(round(y1))
        r0 = max(0, r0); r1 = min(data.shape[0]-1, r1)
        sel[0], sel[1] = r0, r1
        if patch[0]: patch[0].remove()
        patch[0] = patches.Rectangle((-0.5, r0), data.shape[1], r1-r0,
                                      lw=1.5, edgecolor='cyan',
                                      facecolor='cyan', alpha=0.2)
        ax.add_patch(patch[0])
        ax.set_title(f'Rows {r0}-{r1} ({r1-r0+1} rows) — close when happy')
        fig.canvas.draw_idle()

    SpanSelector(ax, onselect, 'vertical', useblit=True,
                 props=dict(facecolor='cyan', alpha=0.25))
    plt.tight_layout(); plt.show()
    return sel[0], sel[1]

# ── Model ─────────────────────────────────────────────────────────────────────

def erf_gauss(x, A_amp, B_amp, C, d, f):
    z = (2 * np.sqrt(np.log(2)) / f) * (d - x)
    return A_amp * (1 + erf(z)) + B_amp * np.exp(-np.log(16)/f**2 * (d-x)**2) + C

# ── 3-5. Collapse, fit, report ────────────────────────────────────────────────

def analyse(data, r0, r1, nm_per_px, h_nm):
    # 3. Collapse
    profile = np.mean(data[r0:r1+1, :], axis=0)
    x_px    = np.arange(len(profile), dtype=float)
    x_nm    = x_px * nm_per_px

    # Edge guess from max gradient
    d_guess = float(x_px[np.argmax(np.abs(np.gradient(profile)))])
    amp     = (profile.max() - profile.min()) / 2

    # 4. Fit Erf + Gaussian
    p0     = [amp, amp*0.3, profile.min(), d_guess, 8.0]
    bounds = ([-np.inf, 0, -np.inf, d_guess-80, 0.5],
              [ np.inf, np.inf, np.inf, d_guess+80, 300.0])
    try:
        popt, _ = curve_fit(erf_gauss, x_px, profile, p0=p0,
                            bounds=bounds, maxfev=10000)
    except Exception as e:
        print(f"Fit failed: {e}"); return

    A_amp, B_amp, C, d_px, f_px = popt
    f_nm  = f_px * nm_per_px
    d_nm  = d_px * nm_per_px

    # 5. Sidewall angle
    theta = 90 - np.degrees(np.arctan(f_nm / h_nm))

    # ── Plot ──────────────────────────────────────────────────────────────────
    fig, axes = plt.subplots(1, 2, figsize=(12, 4.5), facecolor=BG)
    for ax in axes: ax.set_facecolor(BG)

    # Left: full profile + edge marker
    axes[0].plot(x_nm, profile, color=B, lw=1.4)
    axes[0].axvline(d_nm, color=R, lw=1.5, ls='--', label=f'edge d = {d_nm:.1f} nm')
    axes[0].set_xlabel('Position (nm)'); axes[0].set_ylabel('Intensity (a.u.)')
    axes[0].set_title(f'Collapsed profile (rows {r0}-{r1})')
    axes[0].legend(fontsize=9); axes[0].grid(True, lw=0.3, alpha=0.4)

    # Right: fit window zoom
    win_px = 80
    lo = max(0, int(d_px)-win_px); hi = min(len(x_px), int(d_px)+win_px)
    xw_nm = x_nm[lo:hi]; yw = profile[lo:hi]
    x_fit = np.linspace(x_px[lo], x_px[hi-1], 500)
    axes[1].scatter(xw_nm, yw, s=14, color=B, alpha=0.7, label='data')
    axes[1].plot(x_fit * nm_per_px,
                 erf_gauss(x_fit, *popt), color=R, lw=2, label='Erf+Gauss fit')
    axes[1].axvline(d_nm, color=R, lw=1, ls='--')
    axes[1].set_xlabel('Position (nm)'); axes[1].set_ylabel('Intensity (a.u.)')
    axes[1].set_title(f'Edge fit   f = {f_nm:.2f} nm   θ = {theta:.2f}°')
    axes[1].legend(fontsize=9); axes[1].grid(True, lw=0.3, alpha=0.4)

    plt.tight_layout()
    out = os.path.splitext(os.path.basename(path))[0] + '_edge_fit.png'
    plt.savefig(out, dpi=150, bbox_inches='tight')
    print(f"Saved: {out}")
    plt.show()

    # ── Results ───────────────────────────────────────────────────────────────
    print("\n" + "─"*44)
    print("Results")
    print("─"*44)
    print(f"  Pixel size        = {nm_per_px} nm/px")
    print(f"  Feature height h  = {h_nm} nm")
    print(f"  Edge position d   = {d_px:.1f} px = {d_nm:.1f} nm")
    print(f"  FWHM f            = {f_px:.2f} px = {f_nm:.2f} nm")
    print(f"  A (Erf amp)       = {A_amp:.3f}")
    print(f"  B (Gauss amp)     = {B_amp:.3f}")
    print(f"  C (baseline)      = {C:.3f}")
    print(f"  Sidewall angle θ  = {theta:.2f}°")
    print(f"  Meets >=89.4°     = {'YES' if theta >= 89.4 else 'NO'}")
    print("─"*44)

# ── Main ──────────────────────────────────────────────────────────────────────

path      = input("Excel file path: ").strip()
nm_per_px = float(input("Pixel size (nm/px): ").strip())
h_nm      = float(input("Feature height h (nm): ").strip())

if not os.path.exists(path):
    print("File not found."); sys.exit(1)

data = load(path)
r0, r1 = pick_rows(data)
if r0 is None:
    r0, r1 = 0, data.shape[0]-1

analyse(data, r0, r1, nm_per_px, h_nm)