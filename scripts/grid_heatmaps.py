"""
Electron detector grid scan viewer
====================================
Loads one or more Excel scan files and displays them as interactive
heatmaps. Supports:
  - Colourmap selection
  - Percentile-based contrast clipping
  - Click to read pixel value
  - Optional Gaussian smoothing to reduce shot noise
  - Row / column profile plots on click

Usage:
    python grid_viewer.py                      # prompts for file(s)
    python grid_viewer.py scan1.xlsx           # single file
    python grid_viewer.py scan1.xlsx scan2.xlsx  # side-by-side comparison
"""

import sys
import os
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.widgets import Slider, Button, RadioButtons
from scipy.ndimage import gaussian_filter
from openpyxl import load_workbook


# ── Load ──────────────────────────────────────────────────────────────────────

def load_xlsx(path):
    print(f"  Loading {os.path.basename(path)} ...", end=" ", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    arr = np.array(
        [[v if isinstance(v, (int, float)) else 0 for v in r] for r in rows],
        dtype=float
    )
    wb.close()
    print(f"done  {arr.shape}  mean={arr.mean():.3f}  max={arr.max():.0f}")
    return arr


# ── Single-file interactive viewer ────────────────────────────────────────────

def view_single(arr, title):
    fig = plt.figure(figsize=(12, 8), facecolor='#f8f8f6')
    fig.canvas.manager.set_window_title(title)

    gs = gridspec.GridSpec(
        3, 3,
        figure=fig,
        left=0.07, right=0.97,
        top=0.93, bottom=0.22,
        hspace=0.45, wspace=0.35
    )

    ax_img  = fig.add_subplot(gs[:2, :2])
    ax_row  = fig.add_subplot(gs[0, 2])
    ax_col  = fig.add_subplot(gs[1, 2])
    ax_hist = fig.add_subplot(gs[2, :])

    ax_img.set_facecolor('#f8f8f6')
    ax_row.set_facecolor('#f8f8f6')
    ax_col.set_facecolor('#f8f8f6')
    ax_hist.set_facecolor('#f8f8f6')

    state = {'sigma': 0.0, 'plo': 0, 'phi': 99, 'cmap': 'hot'}

    def get_display():
        d = arr.copy()
        if state['sigma'] > 0:
            d = gaussian_filter(d, sigma=state['sigma'])
        vmin = np.percentile(d, state['plo'])
        vmax = np.percentile(d, state['phi'])
        return d, vmin, vmax

    d, vmin, vmax = get_display()
    im = ax_img.imshow(d, cmap=state['cmap'], aspect='equal',
                       vmin=vmin, vmax=vmax, interpolation='nearest')
    cb = fig.colorbar(im, ax=ax_img, label='Electron count', shrink=0.9)
    ax_img.set_title(title, fontsize=10, fontweight='500')
    ax_img.set_xlabel('Column (px)', fontsize=9)
    ax_img.set_ylabel('Row (px)', fontsize=9)

    # Profile lines (invisible until click)
    hline = ax_img.axhline(-1, color='#185FA5', lw=0.8, ls='--', alpha=0.7)
    vline = ax_img.axvline(-1, color='#0F6E56', lw=0.8, ls='--', alpha=0.7)
    dot   = ax_img.plot([], [], 'r+', ms=12, mew=1.5)[0]

    row_line, = ax_row.plot([], [], color='#185FA5', lw=1.2)
    col_line, = ax_col.plot([], [], color='#0F6E56', lw=1.2)

    ax_row.set_title('Row profile', fontsize=9)
    ax_row.set_xlabel('Column (px)', fontsize=8)
    ax_row.set_ylabel('Count', fontsize=8)
    ax_row.tick_params(labelsize=7)
    ax_row.grid(True, lw=0.3, alpha=0.4)

    ax_col.set_title('Column profile', fontsize=9)
    ax_col.set_xlabel('Row (px)', fontsize=8)
    ax_col.set_ylabel('Count', fontsize=8)
    ax_col.tick_params(labelsize=7)
    ax_col.grid(True, lw=0.3, alpha=0.4)

    # Histogram
    flat = arr.flatten()
    ax_hist.hist(flat, bins=min(int(arr.max()) + 1, 50),
                 color='#185FA5', alpha=0.7, edgecolor='none')
    ax_hist.set_xlabel('Electron count', fontsize=9)
    ax_hist.set_ylabel('Frequency', fontsize=9)
    ax_hist.set_title('Count distribution', fontsize=9)
    ax_hist.tick_params(labelsize=8)
    ax_hist.grid(True, lw=0.3, alpha=0.4)

    # ── Widgets ───────────────────────────────────────────────────────────────
    ax_sig  = fig.add_axes([0.10, 0.13, 0.25, 0.03])
    ax_plo  = fig.add_axes([0.10, 0.09, 0.25, 0.03])
    ax_phi  = fig.add_axes([0.10, 0.05, 0.25, 0.03])
    ax_cmap = fig.add_axes([0.55, 0.02, 0.18, 0.14])
    ax_rst  = fig.add_axes([0.80, 0.05, 0.10, 0.06])

    sl_sig = Slider(ax_sig, 'Smooth σ', 0, 5, valinit=0, valstep=0.5,
                    color='#185FA5')
    sl_plo = Slider(ax_plo, 'Clip lo %', 0, 50, valinit=0, valstep=1,
                    color='#888780')
    sl_phi = Slider(ax_phi, 'Clip hi %', 50, 100, valinit=99, valstep=1,
                    color='#888780')
    rb_cmap = RadioButtons(ax_cmap, ['hot', 'viridis', 'cividis', 'gray', 'plasma'],
                           active=0)
    btn_rst = Button(ax_rst, 'Reset', color='#f1efe8')

    def redraw():
        d, vmin, vmax = get_display()
        im.set_data(d)
        im.set_clim(vmin, vmax)
        im.set_cmap(state['cmap'])
        fig.canvas.draw_idle()

    def on_sigma(val):
        state['sigma'] = val
        redraw()

    def on_plo(val):
        state['plo'] = val
        redraw()

    def on_phi(val):
        state['phi'] = val
        redraw()

    def on_cmap(label):
        state['cmap'] = label
        redraw()

    def on_reset(event):
        sl_sig.set_val(0)
        sl_plo.set_val(0)
        sl_phi.set_val(99)
        state.update({'sigma': 0, 'plo': 0, 'phi': 99, 'cmap': 'hot'})
        redraw()

    sl_sig.on_changed(on_sigma)
    sl_plo.on_changed(on_plo)
    sl_phi.on_changed(on_phi)
    rb_cmap.on_clicked(on_cmap)
    btn_rst.on_clicked(on_reset)

    # ── Click handler ─────────────────────────────────────────────────────────
    def on_click(event):
        if event.inaxes != ax_img:
            return
        c, r = int(round(event.xdata)), int(round(event.ydata))
        c = max(0, min(arr.shape[1] - 1, c))
        r = max(0, min(arr.shape[0] - 1, r))

        hline.set_ydata([r, r])
        vline.set_xdata([c, c])
        dot.set_data([c], [r])

        row_data = arr[r, :]
        col_data = arr[:, c]
        row_line.set_data(np.arange(len(row_data)), row_data)
        col_line.set_data(np.arange(len(col_data)), col_data)

        ax_row.set_xlim(0, len(row_data))
        ax_row.set_ylim(0, row_data.max() + 1)
        ax_row.set_title(f'Row {r}  (col {c} val={arr[r,c]:.0f})', fontsize=9)

        ax_col.set_xlim(0, len(col_data))
        ax_col.set_ylim(0, col_data.max() + 1)
        ax_col.set_title(f'Col {c}  (row {r} val={arr[r,c]:.0f})', fontsize=9)

        fig.canvas.draw_idle()

    fig.canvas.mpl_connect('button_press_event', on_click)
    plt.show()


# ── Side-by-side comparison viewer ───────────────────────────────────────────

def view_comparison(arrays, titles):
    n = len(arrays)
    fig, axes = plt.subplots(1, n, figsize=(6 * n, 6), facecolor='#f8f8f6')
    if n == 1:
        axes = [axes]
    fig.subplots_adjust(wspace=0.35, bottom=0.18)

    ims = []
    for ax, arr, title in zip(axes, arrays, titles):
        ax.set_facecolor('#f8f8f6')
        vmax = np.percentile(arr, 99)
        im = ax.imshow(arr, cmap='hot', aspect='equal',
                       vmin=0, vmax=vmax if vmax > 0 else 1,
                       interpolation='nearest')
        plt.colorbar(im, ax=ax, label='count', shrink=0.85)
        ax.set_title(f'{title}\n{arr.shape[0]}x{arr.shape[1]}'
                     f'  mean={arr.mean():.2f}  max={arr.max():.0f}',
                     fontsize=9, fontweight='500')
        ax.set_xlabel('Column (px)', fontsize=8)
        ax.set_ylabel('Row (px)', fontsize=8)
        ax.tick_params(labelsize=7)
        ims.append(im)

    # Shared percentile clip slider
    ax_phi = fig.add_axes([0.2, 0.05, 0.6, 0.03])
    sl_phi = Slider(ax_phi, 'Clip hi %', 80, 100, valinit=99, valstep=1,
                    color='#185FA5')

    def on_phi(val):
        for arr, im in zip(arrays, ims):
            im.set_clim(0, np.percentile(arr, val))
        fig.canvas.draw_idle()

    sl_phi.on_changed(on_phi)
    plt.show()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    while True:
        path = input("\nEnter Excel file path (or 'q' to quit):\n> ").strip()
        if path.lower() == 'q':
            print("Exiting.")
            break
        if not os.path.exists(path):
            print(f"File not found: {path}")
            continue
        if not (path.endswith('.xlsx') or path.endswith('.xls')):
            print("Please provide an .xlsx file.")
            continue
        arr   = load_xlsx(path)
        title = os.path.basename(path).replace('.xlsx', '')
        print("Click on the heatmap to see row/column profiles.")
        view_single(arr, title)
 
 
if __name__ == '__main__':
    main()

if __name__ == '__main__':
    main()