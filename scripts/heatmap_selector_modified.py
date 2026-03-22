import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl import load_workbook
from matplotlib.widgets import RectangleSelector, Button
from scipy.optimize import curve_fit
from scipy.special import erf
from matplotlib.gridspec import GridSpec

# Global variables
df_global = None
ax_global = None
fig_global = None
rect_selector = None
selections = []  # Store multiple selections
selection_mode = 'single'  # 'single', 'multi', or 'rowsets'
pixel_size_x_nm = None  # Pixel size in X direction (nm)
pixel_size_y_nm = None  # Pixel size in Y direction (nm)

class Selection:
    """Store information about a selected region"""
    def __init__(self, name, row_range, col_range, data, total_counts):
        self.name = name
        self.row_range = row_range
        self.col_range = col_range
        self.data = data
        self.total_counts = total_counts

def pixels_to_nm_x(pixel_indices):
    """Convert pixel indices to nanometers in X direction"""
    global pixel_size_x_nm
    if pixel_size_x_nm is None:
        return pixel_indices
    return pixel_indices * pixel_size_x_nm

def pixels_to_nm_y(pixel_indices):
    """Convert pixel indices to nanometers in Y direction"""
    global pixel_size_y_nm
    if pixel_size_y_nm is None:
        return pixel_indices
    return pixel_indices * pixel_size_y_nm

def edge_function(x, A, B, C, d, f):
    """
    Edge fitting function
    F = A[1 + Erf(2√(ln2)/f * (d-x))] + B×exp[-ln16/f² * (d-x)²] + C
    
    Parameters:
    x: scanning position
    A: amplitude of error function component
    B: amplitude of Gaussian peak  
    C: baseline signal value
    d: physical position of sharp edge
    f: FWHM of the beam spot
    """
    ln2 = np.log(2)
    ln16 = np.log(16)
    
    erf_term = A * (1 + erf((2 * np.sqrt(ln2) / f) * (d - x)))
    gaussian_term = B * np.exp(-(ln16 / (f**2)) * (d - x)**2)
    
    return erf_term + gaussian_term + C

def fit_curve(x_data, y_data):
    """
    Fit the edge function to data
    Returns fitted parameters and the fitted curve
    """
    try:
        # Initial parameter guesses
        C_guess = np.min(y_data)  # Baseline
        A_guess = (np.max(y_data) - np.min(y_data)) / 2  # Error function amplitude
        B_guess = (np.max(y_data) - np.min(y_data)) / 4  # Gaussian amplitude
        d_guess = x_data[np.argmax(np.gradient(y_data))]  # Edge position
        f_guess = (x_data.max() - x_data.min()) / 10  # FWHM guess
        
        initial_guess = [A_guess, B_guess, C_guess, d_guess, f_guess]
        
        # Set reasonable bounds
        bounds = (
            [0, 0, 0, x_data.min(), 0.1],  # Lower bounds
            [np.inf, np.inf, np.max(y_data), x_data.max(), x_data.max() - x_data.min()]  # Upper bounds
        )
        
        # Fit the curve
        popt, pcov = curve_fit(edge_function, x_data, y_data, 
                              p0=initial_guess, bounds=bounds, maxfev=10000)
        
        return popt, pcov
    
    except Exception as e:
        print("Curve fitting failed:", e)
        return None, None

def onselect(eclick, erelease):
    """Handle rectangle selection on heatmap"""
    global df_global, selection_mode, selections
    
    x1, y1 = eclick.xdata, eclick.ydata
    x2, y2 = erelease.xdata, erelease.ydata
    
    # Ensure coordinates are in correct order
    x_min, x_max = min(x1, x2), max(x1, x2)
    y_min, y_max = min(y1, y2), max(y1, y2)
    
    # Convert to integer indices
    col_min, col_max = int(np.floor(x_min)), int(np.ceil(x_max))
    row_min, row_max = int(np.floor(y_min)), int(np.ceil(y_max))
    
    print("\nSelected rectangle: (", x_min, ",", y_min, ") to (", x_max, ",", y_max, ")")
    print("Row indices:", row_min, "to", row_max)
    print("Column indices:", col_min, "to", col_max)
    
    # Extract selected region
    selected_data = df_global.iloc[row_min:row_max, col_min:col_max]
    
    if selection_mode == 'single':
        # Single selection behavior with line fitting - SUM instead of average
        col_sums = selected_data.sum(axis=0).values
        col_indices = np.arange(col_min, col_max)
        plot_line_fit(col_indices, col_sums, col_min, col_max)
        
    elif selection_mode == 'multi':
        # Multiple selection comparison mode - SUM instead of average
        total_counts = selected_data.values.sum()
        selection_name = "Selection " + str(len(selections) + 1)
        
        selection = Selection(
            name=selection_name,
            row_range=(row_min, row_max),
            col_range=(col_min, col_max),
            data=selected_data,
            total_counts=total_counts
        )
        selections.append(selection)
        
        print("\n" + selection_name + " - Total counts:", total_counts)
        print("Total selections:", len(selections))
        
        # Plot comparison
        plot_multi_selection_comparison()
        
    elif selection_mode == 'rowsets':
        # Row-wise set analysis mode
        handle_rowset_selection(selected_data, row_min, row_max, col_min, col_max)

def plot_line_fit(x_data, y_data, col_min, col_max):
    """Plot summed values with edge function fit"""
    
    fig, ax1 = plt.subplots(1, 1, figsize=(12, 8))
    
    # Convert x_data to nanometers
    x_data_nm = pixels_to_nm_x(x_data)
    
    # Plot data with line fit
    ax1.plot(x_data_nm, y_data, 'o', markersize=8, label='Summed Data Points', color='blue')
    
    if len(x_data) < 5:
        print("Not enough points for line fitting (need at least 5)")
        if pixel_size_x_nm is not None:
            ax1.set_xlabel('Nanometers', fontsize=12)
        else:
            ax1.set_xlabel('Column Index', fontsize=12)
        ax1.set_ylabel('Counts', fontsize=12)
        ax1.set_title('Column-wise Sums', fontsize=14)
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.show()
        return
    
    # Fit curve (using nm data)
    popt, pcov = fit_curve(x_data_nm, y_data)
    
    if popt is not None:
        A, B, C, d, f = popt
        
        # Generate smooth curve
        x_smooth = np.linspace(x_data_nm.min(), x_data_nm.max(), 200)
        y_smooth = edge_function(x_smooth, A, B, C, d, f)
        
        # Plot fitted curve
        ax1.plot(x_smooth, y_smooth, '-', linewidth=2, label='Edge Fit', color='red')
        
        # Print parameters separately (no text box on plot)
        print("\n" + "="*50)
        print("FITTED PARAMETERS")
        print("="*50)
        print("A (Erf amplitude)    : {:.4f}".format(A))
        print("B (Gaussian amplitude): {:.4f}".format(B))
        print("C (Baseline)         : {:.4f}".format(C))
        if pixel_size_x_nm is not None:
            print("d (Edge position)    : {:.4f} nm".format(d))
            print("f (FWHM)             : {:.4f} nm".format(f))
        else:
            print("d (Edge position)    : {:.4f}".format(d))
            print("f (FWHM)             : {:.4f}".format(f))
        print("="*50)
        
    else:
        print("Line fitting failed, showing data only")
    
    # Formatting
    if pixel_size_x_nm is not None:
        ax1.set_xlabel('Nanometers', fontsize=12)
    else:
        ax1.set_xlabel('Column Index', fontsize=12)
    ax1.set_ylabel('Counts', fontsize=12)
    ax1.set_title('Edge Line Fit - Columns ' + str(col_min) + '-' + str(col_max), fontsize=14)
    ax1.legend(loc='upper right')
    ax1.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.show()

def plot_multi_selection_comparison():
    """Plot comparison of multiple selections"""
    if len(selections) == 0:
        return
    
    fig = plt.figure(figsize=(14, 8))
    gs = GridSpec(2, 2, figure=fig, hspace=0.3, wspace=0.3)
    
    # Plot 1: Bar chart of total counts
    ax1 = fig.add_subplot(gs[0, 0])
    names = [s.name for s in selections]
    totals = [s.total_counts for s in selections]
    colors = plt.cm.Set3(np.linspace(0, 1, len(selections)))
    
    bars = ax1.bar(names, totals, color=colors, edgecolor='black', linewidth=1.5)
    ax1.set_ylabel('Total Counts', fontsize=12)
    ax1.set_title('Total Counts Comparison', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3, axis='y')
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    # Add value labels on bars
    for bar, total in zip(bars, totals):
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height,
                '{:.0f}'.format(total),
                ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    # Plot 2: Heatmap comparison of all selections
    ax2 = fig.add_subplot(gs[0, 1])
    
    # Create comparison matrix
    max_rows = max([s.data.shape[0] for s in selections])
    max_cols = max([s.data.shape[1] for s in selections])
    
    comparison_data = []
    for s in selections:
        # Pad data to match dimensions
        padded = np.zeros((max_rows, max_cols))
        padded[:s.data.shape[0], :s.data.shape[1]] = s.data.values
        comparison_data.append(padded.flatten())
    
    comparison_matrix = np.array(comparison_data)
    
    sns.heatmap(comparison_matrix, ax=ax2, cmap='viridis', 
                yticklabels=names, cbar_kws={'label': 'Counts'})
    ax2.set_title('Selection Heatmaps', fontsize=14, fontweight='bold')
    ax2.set_xlabel('Position Index', fontsize=12)
    
    # Plot 3: Individual selection details
    ax3 = fig.add_subplot(gs[1, :])
    
    details_text = "SELECTION DETAILS:\n\n"
    for s in selections:
        details_text += "{}\n".format(s.name)
        details_text += "  Rows: {} - {}\n".format(s.row_range[0], s.row_range[1])
        details_text += "  Cols: {} - {}\n".format(s.col_range[0], s.col_range[1])
        details_text += "  Shape: {} x {}\n".format(s.data.shape[0], s.data.shape[1])
        details_text += "  Total Counts: {:.2f}\n".format(s.total_counts)
        details_text += "  Max Value: {:.2f}\n".format(s.data.values.max())
        details_text += "  Min Value: {:.2f}\n\n".format(s.data.values.min())
    
    ax3.text(0.05, 0.95, details_text, transform=ax3.transAxes,
            fontsize=9, verticalalignment='top', fontfamily='monospace',
            bbox=dict(boxstyle='round', facecolor='lightgray', alpha=0.8))
    ax3.axis('off')
    
    plt.suptitle('Multi-Selection Comparison ({} selections)'.format(len(selections)),
                fontsize=16, fontweight='bold')
    plt.tight_layout()
    plt.show()

def handle_rowset_selection(selected_data, row_min, row_max, col_min, col_max):
    """Handle row-wise set analysis"""
    total_rows = row_max - row_min
    
    print("\nTotal rows selected:", total_rows)
    rows_per_set = int(input("Enter number of rows per set: "))
    
    if rows_per_set < 1 or rows_per_set > total_rows:
        print("Invalid rows per set. Must be between 1 and", total_rows)
        return
    
    num_sets = total_rows // rows_per_set
    remaining_rows = total_rows % rows_per_set
    
    print("Number of complete sets:", num_sets)
    if remaining_rows > 0:
        print("Remaining rows (will be ignored):", remaining_rows)
    
    # Process each set - SUM instead of average
    set_results = []
    for i in range(num_sets):
        start_row = i * rows_per_set
        end_row = start_row + rows_per_set
        
        set_data = selected_data.iloc[start_row:end_row, :]
        col_sums = set_data.sum(axis=0).values  # Changed from mean to sum
        col_indices = np.arange(col_min, col_max)
        col_indices_nm = pixels_to_nm_x(col_indices)
        
        # Fit curve
        popt, pcov = fit_curve(col_indices_nm, col_sums)
        
        set_results.append({
            'set_num': i + 1,
            'row_range': (row_min + start_row, row_min + end_row),
            'x_data': col_indices_nm,
            'y_data': col_sums,
            'params': popt
        })
        
        if popt is not None:
            A, B, C, d, f = popt
            print("\nSet {} - Fitted Parameters:".format(i + 1))
            print("  A (Erf amplitude)    : {:.4f}".format(A))
            print("  B (Gaussian amplitude): {:.4f}".format(B))
            print("  C (Baseline)         : {:.4f}".format(C))
            if pixel_size_x_nm is not None:
                print("  d (Edge position)    : {:.4f} nm".format(d))
                print("  f (FWHM)             : {:.4f} nm".format(f))
            else:
                print("  d (Edge position)    : {:.4f}".format(d))
                print("  f (FWHM)             : {:.4f}".format(f))
    
    # Create multi-panel plot
    ncols = min(3, num_sets)
    nrows = (num_sets + ncols - 1) // ncols
    
    fig, axes = plt.subplots(nrows, ncols, figsize=(5*ncols, 4*nrows))
    axes = axes.flatten() if num_sets > 1 else [axes]
    
    colors = plt.cm.rainbow(np.linspace(0, 1, num_sets))
    
    for idx, result in enumerate(set_results):
        ax = axes[idx]
        x_data = result['x_data']
        y_data = result['y_data']
        popt = result['params']
        
        # Plot data points
        ax.plot(x_data, y_data, 'o', markersize=6, color=colors[idx], 
               label='Set ' + str(result["set_num"]), alpha=0.7)
        
        # Plot fit if available
        if popt is not None:
            x_smooth = np.linspace(x_data.min(), x_data.max(), 200)
            y_smooth = edge_function(x_smooth, *popt)
            ax.plot(x_smooth, y_smooth, '-', linewidth=2, color='black', alpha=0.8)
        
        ax.set_title('Set ' + str(result["set_num"]) + ' (Rows ' + str(result["row_range"][0]) + '-' + str(result["row_range"][1]) + ')',
                    fontsize=10, fontweight='bold')
        if pixel_size_x_nm is not None:
            ax.set_xlabel('Nanometers', fontsize=9)
        else:
            ax.set_xlabel('Column Index', fontsize=9)
        ax.set_ylabel('Counts', fontsize=9)
        ax.grid(True, alpha=0.3)
    
    # Hide unused subplots
    for idx in range(num_sets, len(axes)):
        axes[idx].axis('off')
    
    plt.suptitle('Row-wise Set Analysis (' + str(num_sets) + ' sets, ' + str(rows_per_set) + ' rows each, Cols ' + str(col_min) + '-' + str(col_max) + ')',
                fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.show()
    
    # Create overlay comparison plot
    fig2, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    # Plot 1: All fits overlaid
    for idx, result in enumerate(set_results):
        x_data = result['x_data']
        y_data = result['y_data']
        popt = result['params']
        
        ax1.plot(x_data, y_data, 'o', markersize=4, color=colors[idx], 
                alpha=0.6, label='Set ' + str(result["set_num"]))
        
        if popt is not None:
            x_smooth = np.linspace(x_data.min(), x_data.max(), 200)
            y_smooth = edge_function(x_smooth, *popt)
            ax1.plot(x_smooth, y_smooth, '-', linewidth=2, color=colors[idx], alpha=0.8)
    
    if pixel_size_x_nm is not None:
        ax1.set_xlabel('Nanometers', fontsize=12)
    else:
        ax1.set_xlabel('Column Index', fontsize=12)
    ax1.set_ylabel('Counts', fontsize=12)
    ax1.set_title('Overlay Comparison - All Sets', fontsize=14, fontweight='bold')
    ax1.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=8)
    ax1.grid(True, alpha=0.3)
    
    # Plot 2: Parameter comparison
    param_names = ['A (Erf)', 'B (Gauss)', 'C (Base)', 'd (Edge)', 'f (FWHM)']
    param_matrix = np.array([r['params'] for r in set_results if r['params'] is not None])
    
    if len(param_matrix) > 0:
        x_pos = np.arange(len(param_names))
        width = 0.8 / num_sets
        
        for idx in range(num_sets):
            if set_results[idx]['params'] is not None:
                offset = (idx - num_sets/2) * width + width/2
                ax2.bar(x_pos + offset, param_matrix[idx], width, 
                       label='Set ' + str(idx+1), color=colors[idx], alpha=0.8)
        
        ax2.set_xticks(x_pos)
        ax2.set_xticklabels(param_names, rotation=45, ha='right')
        ax2.set_ylabel('Parameter Value', fontsize=12)
        ax2.set_title('Parameter Comparison Across Sets', fontsize=14, fontweight='bold')
        ax2.legend(fontsize=8)
        ax2.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    plt.show()

def visualize_excel_counts(file_path, colourmap, figsize=(12, 10), show_values=False):
    """Create heatmap visualization with nanometer axes starting from same origin"""
    global df_global, ax_global, fig_global, pixel_size_x_nm, pixel_size_y_nm
    
    filename = file_path.split("/")[-1]
    title = filename
    wb = load_workbook(file_path)
    sheet_name = wb.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    df = df.dropna(how='all').dropna(axis=1, how='all')
    df = df.fillna(0)
    df = df.apply(pd.to_numeric, errors='coerce').fillna(0)
    
    df_global = df
    
    print("Data shape:", df.shape)
    print("Value range:", df.values.min(), "to", df.values.max())
    
    fig_global = plt.figure(figsize=figsize)
    ax = sns.heatmap(df, cmap=colourmap, annot=show_values, fmt='.0f', 
                     cbar_kws={'label': 'Counts'}, square=False)
    
    ax_global = ax
    
    # Convert axis labels to nanometers, starting from same origin (0,0)
    if pixel_size_x_nm is not None and pixel_size_y_nm is not None:
        # Set x-axis labels in nanometers starting from 0
        num_xticks = min(10, df.shape[1])
        x_positions = np.linspace(0, df.shape[1], num_xticks)
        x_labels = ['{:.1f}'.format(pixels_to_nm_x(pos)) for pos in x_positions]
        ax.set_xticks(x_positions)
        ax.set_xticklabels(x_labels, rotation=0)
        
        # Set y-axis labels in nanometers starting from 0
        num_yticks = min(10, df.shape[0])
        y_positions = np.linspace(0, df.shape[0], num_yticks)
        y_labels = ['{:.1f}'.format(pixels_to_nm_y(pos)) for pos in y_positions]
        ax.set_yticks(y_positions)
        ax.set_yticklabels(y_labels, rotation=1)
        # ax.yaxis.set_inverted(True)
        plt.xlabel('Nanometers', fontsize=12)
        plt.ylabel('Nanometers', fontsize=12)
    else:
        plt.xlabel('Column', fontsize=12)
        plt.ylabel('Row', fontsize=12)
    
    plt.title(title, fontsize=16, pad=20)
    plt.tight_layout()
    
    return plt, df, ax

def check_file(file_path):
    """Check if file exists and is valid"""
    try:
        wb = load_workbook(file_path)
        return True
    except:
        print("No such file, please re-enter")
        return False

def print_menu():
    """Print mode selection menu"""
    print("\n" + "="*60)
    print("HEATMAP SELECTOR - MODE SELECTION")
    print("="*60)
    print("1. Single Selection (Line Fit)")
    print("   - Select one region at a time")
    print("   - Fit edge function to column sums")
    print()
    print("2. Multi-Selection Comparison")
    print("   - Select multiple regions")
    print("   - Compare total counts across selections")
    print()
    print("3. Row-wise Set Analysis")
    print("   - Select region and divide into row sets")
    print("   - Fit each set separately for comparison")
    print()
    print("4. Exit")
    print("="*60)

def user_input():
    """Main user interaction function"""
    global selection_mode, rect_selector, selections, pixel_size_x_nm, pixel_size_y_nm
    
    # Get file path
    file_correct = False
    while not file_correct:
        file_path = input("Enter filepath and name: ")
        file_correct = check_file(file_path)
    
    # Get pixel size parameters
    print("\n" + "="*60)
    print("PIXEL SIZE CONFIGURATION")
    print("="*60)
    print("Each pixel is a cell in your data array.")
    print("Please specify the physical size of each pixel in nanometers.")
    print()
    
    try:
        pixel_x_input = input("Enter pixel size in X direction (nm) [or press Enter to skip]: ")
        if pixel_x_input.strip():
            pixel_size_x_nm = float(pixel_x_input)
            pixel_y_input = input("Enter pixel size in Y direction (nm) [or press Enter to use X value]: ")
            if pixel_y_input.strip():
                pixel_size_y_nm = float(pixel_y_input)
            else:
                pixel_size_y_nm = pixel_size_x_nm
                print("Using same pixel size for Y direction: {:.4f} nm".format(pixel_size_y_nm))
            
            print("\nPixel size configuration:")
            print("  X direction: {:.4f} nm/pixel".format(pixel_size_x_nm))
            print("  Y direction: {:.4f} nm/pixel".format(pixel_size_y_nm))
        else:
            pixel_size_x_nm = None
            pixel_size_y_nm = None
            print("Pixel size not provided - axes will be in pixel units")
    except ValueError:
        print("Invalid pixel size. Using pixel units.")
        pixel_size_x_nm = None
        pixel_size_y_nm = None
    
    print("="*60)
    
    # Get colormap
    colour_map = input("\nChoose a colour scheme ('viridis', 'plasma', 'hot', 'coolwarm', 'Blues'): ")
    if not colour_map:
        colour_map = 'viridis'
    
    while True:
        print_menu()
        choice = input("Select mode (1-4): ")
        
        if choice == '4':
            print("Exiting...")
            break
        
        if choice not in ['1', '2', '3']:
            print("Invalid choice. Please enter 1, 2, 3, or 4.")
            continue
        
        # Set mode
        if choice == '1':
            selection_mode = 'single'
            print("\n=== SINGLE SELECTION MODE ===")
            print("Click and drag to select a region")
            print("Column sums will be fitted with edge function")
        elif choice == '2':
            selection_mode = 'multi'
            selections = []  # Reset selections
            print("\n=== MULTI-SELECTION MODE ===")
            print("Click and drag to select multiple regions")
            print("Each selection will be added to comparison")
            print("Close the plot window to finish and see comparison")
        elif choice == '3':
            selection_mode = 'rowsets'
            print("\n=== ROW-WISE SET ANALYSIS MODE ===")
            print("Click and drag to select a region")
            print("You'll be asked to specify rows per set")
        
        # Create visualization
        plt_obj, df, ax = visualize_excel_counts(file_path, colour_map)
        
        # Set up rectangle selector
        rect_selector = RectangleSelector(ax, onselect, useblit=True, button=[1],
                                         minspanx=5, minspany=5, spancoords='data',
                                         use_data_coordinates=True,
                                         interactive=True)
        
        plt.show()
        
        # After closing the plot
        if selection_mode == 'multi' and len(selections) > 0:
            print("\nCollected", len(selections), "selections")
            response = input("View final comparison? (y/n): ")
            if response.lower() == 'y':
                plot_multi_selection_comparison()

if __name__ == "__main__":
    user_input()
