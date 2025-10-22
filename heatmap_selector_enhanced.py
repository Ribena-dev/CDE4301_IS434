import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl import load_workbook
from matplotlib.widgets import RectangleSelector, Button
from scipy.optimize import curve_fit
from scipy.special import erf
from matplotlib.gridspec import GridSpec

# Global variables
df_global = None
ax_global = None
fig_global = None
rect_selector = None
selections = []  # Store multiple selections
selection_mode = 'single'  # 'single', 'multi', or 'rowsets'

class Selection:
    """Store information about a selected region"""
    def __init__(self, name, row_range, col_range, data, avg_counts):
        self.name = name
        self.row_range = row_range
        self.col_range = col_range
        self.data = data
        self.avg_counts = avg_counts

def edge_function(x, A, B, C, d, f):
    """
    Edge fitting function
    F = A[1 + Erf(2√(ln2)/f * (d-x))] + B×exp[-ln16/f² * (d-x)²] + C
    
    Parameters:
    x: scanning position
    A: amplitude of error function component
    B: amplitude of Gaussian peak  
    C: baseline signal value
    d: physical position of sharp edge
    f: FWHM of the beam spot
    """
    ln2 = np.log(2)
    ln16 = np.log(16)
    
    erf_term = A * (1 + erf((2 * np.sqrt(ln2) / f) * (d - x)))
    gaussian_term = B * np.exp(-(ln16 / (f**2)) * (d - x)**2)
    
    return erf_term + gaussian_term + C

def fit_curve(x_data, y_data):
    """
    Fit the edge function to data
    Returns fitted parameters and the fitted curve
    """
    try:
        # Initial parameter guesses
        C_guess = np.min(y_data)  # Baseline
        A_guess = (np.max(y_data) - np.min(y_data)) / 2  # Error function amplitude
        B_guess = (np.max(y_data) - np.min(y_data)) / 4  # Gaussian amplitude
        d_guess = x_data[np.argmax(np.gradient(y_data))]  # Edge position
        f_guess = (x_data.max() - x_data.min()) / 10  # FWHM guess
        
        initial_guess = [A_guess, B_guess, C_guess, d_guess, f_guess]
        
        # Set reasonable bounds
        bounds = (
            [0, 0, 0, x_data.min(), 0.1],  # Lower bounds
            [np.inf, np.inf, np.max(y_data), x_data.max(), x_data.max() - x_data.min()]  # Upper bounds
        )
        
        # Fit the curve
        popt, pcov = curve_fit(edge_function, x_data, y_data, 
                              p0=initial_guess, bounds=bounds, maxfev=10000)
        
        return popt, pcov
    
    except Exception as e:
        print("Curve fitting failed:", e)
        return None, None

def onselect(eclick, erelease):
    """Handle rectangle selection on heatmap"""
    global df_global, selection_mode, selections
    
    x1, y1 = eclick.xdata, eclick.ydata
    x2, y2 = erelease.xdata, erelease.ydata
    
    # Ensure coordinates are in correct order
    x_min, x_max = min(x1, x2), max(x1, x2)
    y_min, y_max = min(y1, y2), max(y1, y2)
    
    # Convert to integer indices
    col_min, col_max = int(np.floor(x_min)), int(np.ceil(x_max))
    row_min, row_max = int(np.floor(y_min)), int(np.ceil(y_max))
    
    print("\nSelected rectangle: (", x_min, ",", y_min, ") to (", x_max, ",", y_max, ")")
    print("Row indices:", row_min, "to", row_max)
    print("Column indices:", col_min, "to", col_max)
    
    # Extract selected region
    selected_data = df_global.iloc[row_min:row_max, col_min:col_max]
    
    if selection_mode == 'single':
        # Original single selection behavior with line fitting
        col_averages = selected_data.mean(axis=0).values
        col_indices = np.arange(col_min, col_max)
        plot_line_fit(col_indices, col_averages, col_min, col_max)
        
    elif selection_mode == 'multi':
        # Multiple selection comparison mode
        avg_counts = selected_data.values.mean()
        selection_name = "Selection " + str(len(selections) + 1)
        
        selection = Selection(
            name=selection_name,
            row_range=(row_min, row_max),
            col_range=(col_min, col_max),
            data=selected_data,
            avg_counts=avg_counts
        )
        selections.append(selection)
        
        print("\n" + selection_name + " - Average counts:", avg_counts)
        print("Total selections:", len(selections))
        
        # Plot comparison
        plot_multi_selection_comparison()
        
    elif selection_mode == 'rowsets':
        # Row-wise set analysis mode
        handle_rowset_selection(selected_data, row_min, row_max, col_min, col_max)

def plot_line_fit(x_data, y_data, col_min, col_max):
    """Plot averaged values with edge function fit"""
    
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 10))
    
    # Plot 1: Data with line fit
    ax1.plot(x_data, y_data, 'o', markersize=8, label='Averaged Data Points', color='blue')
    
    if len(x_data) < 5:
        print("Not enough points for line fitting (need at least 5)")
        ax1.set_xlabel('Column Index', fontsize=12)
        ax1.set_ylabel('Average Value', fontsize=12)
        ax1.set_title('Column-wise Averages', fontsize=14)
        ax1.legend()
        ax1.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.show()
        return
    
    # Fit curve
    popt, pcov = fit_curve(x_data, y_data)
    
    if popt is not None:
        A, B, C, d, f = popt
        
        # Generate smooth curve
        x_smooth = np.linspace(x_data.min(), x_data.max(), 200)
        y_smooth = edge_function(x_smooth, A, B, C, d, f)
        
        # Plot fitted curve
        ax1.plot(x_smooth, y_smooth, '-', linewidth=2, label='Edge Fit', color='red')
        
        # Create parameter display
        param_text = 'Fitted Parameters:\n'
        param_text += 'A (Erf amplitude) = ' + str(round(A, 4)) + '\n'
        param_text += 'B (Gaussian amplitude) = ' + str(round(B, 4)) + '\n'
        param_text += 'C (Baseline) = ' + str(round(C, 4)) + '\n'
        param_text += 'd (Edge position) = ' + str(round(d, 4)) + '\n'
        param_text += 'f (FWHM) = ' + str(round(f, 4))
        
        ax1.text(0.02, 0.98, param_text, transform=ax1.transAxes, 
                fontsize=9, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8),
                family='monospace')
        
        # Calculate residuals
        y_fit = edge_function(x_data, A, B, C, d, f)
        residuals = y_data - y_fit
        
        # Plot 2: Residuals
        ax2.plot(x_data, residuals, 'o-', markersize=6, color='green')
        ax2.axhline(y=0, color='black', linestyle='--', linewidth=1)
        ax2.set_xlabel('Column Index', fontsize=12)
        ax2.set_ylabel('Residuals', fontsize=12)
        ax2.set_title('Fit Residuals', fontsize=12)
        ax2.grid(True, alpha=0.3)
        
        # Calculate R-squared
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((y_data - np.mean(y_data))**2)
        r_squared = 1 - (ss_res / ss_tot)
        
        ax2.text(0.02, 0.98, 'R² = ' + str(round(r_squared, 6)), transform=ax2.transAxes,
                fontsize=10, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='lightblue', alpha=0.8))
        
        print("\n=== Fitted Parameters ===")
        print("A (Erf amplitude):", A)
        print("B (Gaussian amplitude):", B)
        print("C (Baseline):", C)
        print("d (Edge position):", d)
        print("f (FWHM):", f)
        print("R²:", r_squared)
        
    else:
        print("Line fitting failed, showing data only")
    
    # Formatting
    ax1.set_xlabel('Column Index', fontsize=12)
    ax1.set_ylabel('Average Value', fontsize=12)
    ax1.set_title('Edge Line Fit - Columns ' + str(col_min) + '-' + str(col_max), fontsize=14)
    ax1.legend(loc='upper right')
    ax1.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.show()

def plot_multi_selection_comparison():
    """Plot comparison of multiple selections"""
    if len(selections) == 0:
        return
    
    fig = plt.figure(figsize=(14, 8))
    gs = GridSpec(2, 2, figure=fig, hspace=0.3, wspace=0.3)
    
    # Plot 1: Bar chart of average counts
    ax1 = fig.add_subplot(gs[0, 0])
    names = [s.name for s in selections]
    avgs = [s.avg_counts for s in selections]
    colors = plt.cm.Set3(np.linspace(0, 1, len(selections)))
    
    bars = ax1.bar(names, avgs, color=colors, edgecolor='black', linewidth=1.5)
    ax1.set_ylabel('Average Counts', fontsize=12)
    ax1.set_title('Average Counts Comparison', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3, axis='y')
    plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')
    
    # Add value labels on bars
    for bar, avg in zip(bars, avgs):
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height,
                str(round(avg, 1)), ha='center', va='bottom', fontsize=9)
    
    # Plot 2: Statistics table
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.axis('tight')
    ax2.axis('off')
    
    table_data = []
    table_data.append(['Selection', 'Avg', 'Std', 'Min', 'Max', 'Size'])
    
    for sel in selections:
        data_flat = sel.data.values.flatten()
        size_str = str(sel.data.shape[0]) + '×' + str(sel.data.shape[1])
        table_data.append([
            sel.name,
            str(round(np.mean(data_flat), 1)),
            str(round(np.std(data_flat), 1)),
            str(round(np.min(data_flat), 1)),
            str(round(np.max(data_flat), 1)),
            size_str
        ])
    
    table = ax2.table(cellText=table_data, cellLoc='center', loc='center',
                     colWidths=[0.2, 0.15, 0.15, 0.15, 0.15, 0.2])
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1, 2)
    
    # Style header row
    for i in range(6):
        table[(0, i)].set_facecolor('#4CAF50')
        table[(0, i)].set_text_props(weight='bold', color='white')
    
    ax2.set_title('Selection Statistics', fontsize=14, fontweight='bold', pad=20)
    
    # Plot 3: Distribution comparison
    ax3 = fig.add_subplot(gs[1, :])
    
    for i, sel in enumerate(selections):
        data_flat = sel.data.values.flatten()
        ax3.hist(data_flat, bins=30, alpha=0.5, label=sel.name, color=colors[i])
    
    ax3.set_xlabel('Count Value', fontsize=12)
    ax3.set_ylabel('Frequency', fontsize=12)
    ax3.set_title('Distribution Comparison', fontsize=14, fontweight='bold')
    ax3.legend()
    ax3.grid(True, alpha=0.3, axis='y')
    
    plt.suptitle('Multi-Selection Analysis (' + str(len(selections)) + ' selections)', 
                fontsize=16, fontweight='bold', y=0.995)
    
    plt.show()
    
    # Print summary
    print("\n" + "="*60)
    print("MULTI-SELECTION SUMMARY")
    print("="*60)
    for sel in selections:
        data_flat = sel.data.values.flatten()
        print("\n" + sel.name + ":")
        print("  Position: Rows", sel.row_range[0], "-", sel.row_range[1], ", Cols", sel.col_range[0], "-", sel.col_range[1])
        print("  Average:", np.mean(data_flat))
        print("  Std Dev:", np.std(data_flat))
        print("  Range:", np.min(data_flat), "-", np.max(data_flat))

def handle_rowset_selection(selected_data, row_min, row_max, col_min, col_max):
    """Handle row-wise set analysis"""
    print("\nSelected area:", selected_data.shape[0], "rows ×", selected_data.shape[1], "columns")
    
    # Get user input for rows per set
    try:
        rows_per_set = int(input("Enter number of rows per set: "))
        if rows_per_set <= 0 or rows_per_set > selected_data.shape[0]:
            print("Invalid input. Must be between 1 and", selected_data.shape[0])
            return
    except ValueError:
        print("Invalid input. Please enter a number.")
        return
    
    num_sets = selected_data.shape[0] // rows_per_set
    if num_sets == 0:
        print("Not enough rows. Need at least", rows_per_set, "rows.")
        return
    
    print("Creating", num_sets, "sets with", rows_per_set, "rows each")
    
    # Process each set
    set_results = []
    
    for i in range(num_sets):
        start_row = i * rows_per_set
        end_row = start_row + rows_per_set
        
        set_data = selected_data.iloc[start_row:end_row, :]
        col_averages = set_data.mean(axis=0).values
        col_indices = np.arange(col_min, col_max)
        
        # Fit curve
        popt, pcov = fit_curve(col_indices, col_averages)
        
        set_results.append({
            'set_num': i + 1,
            'row_range': (row_min + start_row, row_min + end_row),
            'x_data': col_indices,
            'y_data': col_averages,
            'params': popt
        })
    
    # Plot all sets for comparison
    plot_rowset_comparison(set_results, rows_per_set, col_min, col_max)

def plot_rowset_comparison(set_results, rows_per_set, col_min, col_max):
    """Plot comparison of all row sets"""
    num_sets = len(set_results)
    
    # Determine subplot layout
    if num_sets <= 4:
        nrows, ncols = 2, 2
    elif num_sets <= 6:
        nrows, ncols = 2, 3
    elif num_sets <= 9:
        nrows, ncols = 3, 3
    elif num_sets <= 12:
        nrows, ncols = 3, 4
    else:
        nrows, ncols = 4, 5
    
    fig, axes = plt.subplots(nrows, ncols, figsize=(5*ncols, 4*nrows))
    axes = axes.flatten() if num_sets > 1 else [axes]
    
    colors = plt.cm.rainbow(np.linspace(0, 1, num_sets))
    
    for idx, result in enumerate(set_results):
        ax = axes[idx]
        x_data = result['x_data']
        y_data = result['y_data']
        popt = result['params']
        
        # Plot data points
        ax.plot(x_data, y_data, 'o', markersize=6, color=colors[idx], 
               label='Set ' + str(result["set_num"]), alpha=0.7)
        
        # Plot fit if available
        if popt is not None:
            x_smooth = np.linspace(x_data.min(), x_data.max(), 200)
            y_smooth = edge_function(x_smooth, *popt)
            ax.plot(x_smooth, y_smooth, '-', linewidth=2, color='black', alpha=0.8)
            
            # Add parameters
            A, B, C, d, f = popt
            param_text = 'd=' + str(round(d, 2)) + '\nf=' + str(round(f, 2))
            ax.text(0.05, 0.95, param_text, transform=ax.transAxes,
                   fontsize=8, verticalalignment='top',
                   bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))
        
        ax.set_title('Set ' + str(result["set_num"]) + ' (Rows ' + str(result["row_range"][0]) + '-' + str(result["row_range"][1]) + ')',
                    fontsize=10, fontweight='bold')
        ax.set_xlabel('Column Index', fontsize=9)
        ax.set_ylabel('Average Value', fontsize=9)
        ax.grid(True, alpha=0.3)
    
    # Hide unused subplots
    for idx in range(num_sets, len(axes)):
        axes[idx].axis('off')
    
    plt.suptitle('Row-wise Set Analysis (' + str(num_sets) + ' sets, ' + str(rows_per_set) + ' rows each, Cols ' + str(col_min) + '-' + str(col_max) + ')',
                fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.show()
    
    # Create overlay comparison plot
    fig2, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 6))
    
    # Plot 1: All fits overlaid
    for idx, result in enumerate(set_results):
        x_data = result['x_data']
        y_data = result['y_data']
        popt = result['params']
        
        ax1.plot(x_data, y_data, 'o', markersize=4, color=colors[idx], 
                alpha=0.6, label='Set ' + str(result["set_num"]))
        
        if popt is not None:
            x_smooth = np.linspace(x_data.min(), x_data.max(), 200)
            y_smooth = edge_function(x_smooth, *popt)
            ax1.plot(x_smooth, y_smooth, '-', linewidth=2, color=colors[idx], alpha=0.8)
    
    ax1.set_xlabel('Column Index', fontsize=12)
    ax1.set_ylabel('Average Value', fontsize=12)
    ax1.set_title('Overlay Comparison - All Sets', fontsize=14, fontweight='bold')
    ax1.legend(bbox_to_anchor=(1.05, 1), loc='upper left', fontsize=8)
    ax1.grid(True, alpha=0.3)
    
    # Plot 2: Parameter comparison
    param_names = ['A (Erf)', 'B (Gauss)', 'C (Base)', 'd (Edge)', 'f (FWHM)']
    param_matrix = np.array([r['params'] for r in set_results if r['params'] is not None])
    
    if len(param_matrix) > 0:
        x_pos = np.arange(len(param_names))
        width = 0.8 / num_sets
        
        for idx in range(num_sets):
            if set_results[idx]['params'] is not None:
                offset = (idx - num_sets/2) * width + width/2
                ax2.bar(x_pos + offset, param_matrix[idx], width, 
                       label='Set ' + str(idx+1), color=colors[idx], alpha=0.8)
        
        ax2.set_xticks(x_pos)
        ax2.set_xticklabels(param_names, rotation=45, ha='right')
        ax2.set_ylabel('Parameter Value', fontsize=12)
        ax2.set_title('Parameter Comparison Across Sets', fontsize=14, fontweight='bold')
        ax2.legend(fontsize=8)
        ax2.grid(True, alpha=0.3, axis='y')
    
    plt.tight_layout()
    plt.show()

def visualize_excel_counts(file_path, colourmap, figsize=(12, 10), show_values=False):
    """Create heatmap visualization"""
    global df_global, ax_global, fig_global
    
    filename = file_path.split("/")[-1]
    title = filename
    wb = load_workbook(file_path)
    sheet_name = wb.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    df = df.dropna(how='all').dropna(axis=1, how='all')
    df = df.fillna(0)
    df = df.apply(pd.to_numeric, errors='coerce').fillna(0)
    
    df_global = df
    
    print("Data shape:", df.shape)
    print("Value range:", df.values.min(), "to", df.values.max())
    
    fig_global = plt.figure(figsize=figsize)
    ax = sns.heatmap(df, cmap=colourmap, annot=show_values, fmt='.0f', 
                     cbar_kws={'label': 'Count'}, square=False)
    
    ax_global = ax
    
    plt.title(title, fontsize=16, pad=20)
    plt.xlabel('Column', fontsize=12)
    plt.ylabel('Row', fontsize=12)
    plt.tight_layout()
    
    return plt, df, ax

def check_file(file_path):
    """Check if file exists and is valid"""
    try:
        wb = load_workbook(file_path)
        return True
    except:
        print("No such file, please re-enter")
        return False

def print_menu():
    """Print mode selection menu"""
    print("\n" + "="*60)
    print("HEATMAP SELECTOR - MODE SELECTION")
    print("="*60)
    print("1. Single Selection (Line Fit)")
    print("   - Select one region at a time")
    print("   - Fit edge function to column averages")
    print()
    print("2. Multi-Selection Comparison")
    print("   - Select multiple regions")
    print("   - Compare average counts across selections")
    print()
    print("3. Row-wise Set Analysis")
    print("   - Select region and divide into row sets")
    print("   - Fit each set separately for comparison")
    print()
    print("4. Exit")
    print("="*60)

def user_input():
    """Main user interaction function"""
    global selection_mode, rect_selector, selections
    
    # Get file path
    file_correct = False
    while not file_correct:
        file_path = input("Enter filepath and name: ")
        file_correct = check_file(file_path)
    
    # Get colormap
    colour_map = input("Choose a colour scheme ('viridis', 'plasma', 'hot', 'coolwarm', 'Blues'): ")
    if not colour_map:
        colour_map = 'viridis'
    
    while True:
        print_menu()
        choice = input("Select mode (1-4): ")
        
        if choice == '4':
            print("Exiting...")
            break
        
        if choice not in ['1', '2', '3']:
            print("Invalid choice. Please enter 1, 2, 3, or 4.")
            continue
        
        # Set mode
        if choice == '1':
            selection_mode = 'single'
            print("\n=== SINGLE SELECTION MODE ===")
            print("Click and drag to select a region")
            print("Column averages will be fitted with edge function")
        elif choice == '2':
            selection_mode = 'multi'
            selections = []  # Reset selections
            print("\n=== MULTI-SELECTION MODE ===")
            print("Click and drag to select multiple regions")
            print("Each selection will be added to comparison")
            print("Close the plot window to finish and see comparison")
        elif choice == '3':
            selection_mode = 'rowsets'
            print("\n=== ROW-WISE SET ANALYSIS MODE ===")
            print("Click and drag to select a region")
            print("You'll be asked to specify rows per set")
        
        # Create visualization
        plt_obj, df, ax = visualize_excel_counts(file_path, colour_map)
        
        # Set up rectangle selector
        rect_selector = RectangleSelector(ax, onselect, useblit=True, button=[1],
                                         minspanx=5, minspany=5, spancoords='data',
                                         use_data_coordinates=True,
                                         interactive=True)
        
        plt.show()
        
        # After closing the plot
        if selection_mode == 'multi' and len(selections) > 0:
            print("\nCollected", len(selections), "selections")
            response = input("View final comparison? (y/n): ")
            if response.lower() == 'y':
                plot_multi_selection_comparison()

if __name__ == "__main__":
    user_input()
