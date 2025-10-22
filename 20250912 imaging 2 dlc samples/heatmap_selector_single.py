import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl import load_workbook
from matplotlib.widgets import RectangleSelector
from scipy.optimize import curve_fit
from scipy.interpolate import UnivariateSpline

# Global variables to store data and axes
df_global = None
ax_global = None

def polynomial_func(x, *coeffs):
    """Generic polynomial function for curve fitting"""
    return sum(c * x**i for i, c in enumerate(coeffs))

def onselect(eclick, erelease):
    """Handle rectangle selection on heatmap"""
    global df_global, ax_global
    
    x1, y1 = eclick.xdata, eclick.ydata
    x2, y2 = erelease.xdata, erelease.ydata
    
    # Ensure coordinates are in correct order
    x_min, x_max = min(x1, x2), max(x1, x2)
    y_min, y_max = min(y1, y2), max(y1, y2)
    
    # Convert to integer indices
    col_min, col_max = int(np.floor(x_min)), int(np.ceil(x_max))
    row_min, row_max = int(np.floor(y_min)), int(np.ceil(y_max))
    
    print(f"\nSelected rectangle coordinates: ({x_min:.2f}, {y_min:.2f}) to ({x_max:.2f}, {y_max:.2f})")
    print(f"Row indices: {row_min} to {row_max}")
    print(f"Column indices: {col_min} to {col_max}")
    
    # Extract selected region from dataframe
    selected_data = df_global.iloc[row_min:row_max, col_min:col_max]
    
    print(f"\nSelected data shape: {selected_data.shape}")
    print("Selected data:")
    print(selected_data)
    
    # Calculate column-wise averages (average down each column)
    col_averages = selected_data.mean(axis=0).values
    col_indices = np.arange(col_min, col_max)
    
    print(f"\nColumn averages: {col_averages}")
    print(f"Column indices: {col_indices}")
    
    # Plot the averaged data with fitted curve
    plot_averaged_data(col_indices, col_averages, col_min, col_max)

def plot_averaged_data(x_data, y_data, col_min, col_max):
    """Plot averaged values and fit a curve"""
    
    # Create new figure for the plot
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Plot original points
    ax.plot(x_data, y_data, 'o', markersize=8, label='Averaged Data Points', color='blue')
    
    # Determine degree of polynomial to fit based on number of points
    n_points = len(x_data)
    
    if n_points < 2:
        print("Not enough points to fit a curve")
        ax.set_xlabel('Column Index', fontsize=12)
        ax.set_ylabel('Average Value', fontsize=12)
        ax.set_title('Column-wise Averages', fontsize=14)
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.show()
        return
    
    # Try polynomial fitting with different degrees
    poly_degree = min(5, n_points - 1)  
    
    try:
        # Fit polynomial
        coefficients = np.polyfit(x_data, y_data, poly_degree)
        poly_func = np.poly1d(coefficients)
        
        # Generate smooth curve
        x_smooth = np.linspace(x_data.min(), x_data.max(), 100)
        y_smooth = poly_func(x_smooth)
        
        # Plot fitted curve
        ax.plot(x_smooth, y_smooth, '-', linewidth=2, label=f'Fitted Curve (degree {poly_degree})', color='red')
        
        # Create equation string
        equation = "y = "
        for i, coef in enumerate(coefficients):
            power = poly_degree - i
            if i > 0:
                equation += " + " if coef >= 0 else " - "
                coef = abs(coef)
            
            if power == 0:
                equation += f"{coef:.4f}"
            elif power == 1:
                equation += f"{coef:.4f}x"
            else:
                equation += f"{coef:.4f}x^{power}"
        
        # Add equation to plot
        ax.text(0.05, 0.95, equation, transform=ax.transAxes, 
                fontsize=10, verticalalignment='top',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
        
        print(f"\nFitted equation: {equation}")
        
    except Exception as e:
        print(f"Error fitting curve: {e}")
    
    # Formatting
    ax.set_xlabel('Column Index', fontsize=12)
    ax.set_ylabel('Average Value', fontsize=12)
    ax.set_title(f'Column-wise Averages (Columns {col_min}-{col_max})', fontsize=14)
    ax.legend()
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.show()

def visualize_excel_counts(file_path, colourmap, figsize=(12, 10), show_values=False):
    global df_global, ax_global
    
    filename = file_path.split("/")[-1]
    title = filename
    wb = load_workbook(file_path)
    sheet_name = wb.sheetnames[0]
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    df = df.dropna(how='all').dropna(axis=1, how='all')
    df = df.fillna(0)
    df = df.apply(pd.to_numeric, errors='coerce').fillna(0)
    
    df_global = df  # Store globally for selection callback
    
    print("Data shape:", df.shape)
    print("Value range:", df.values.min(), "to", df.values.max())
    
    plt.figure(figsize=figsize)
    ax = sns.heatmap(df, cmap=colourmap, annot=show_values, fmt='.0f', 
                     cbar_kws={'label': 'Count'}, square=False)
    
    ax_global = ax  # Store globally
    
    plt.title(title, fontsize=16, pad=20)
    plt.xlabel('Column', fontsize=12)
    plt.ylabel('Row', fontsize=12)
    plt.tight_layout()
    
    return plt, df, ax

def check_file(file_path):
    try:
        wb = load_workbook(file_path)
        return True
    except:
        print("No such file, please re-enter")
        return False

def user_input():
    file_correct = False
    while file_correct == False:
        file_path = input("Enter filepath and name (in reference to this directory): ")
        file_correct = check_file(file_path)
    
    colour_map = input("Choose a colour scheme ('viridis', 'plasma', 'hot', 'coolwarm', 'Blues'): ")
    plt, df, ax = visualize_excel_counts(file_path, colour_map)
    
    print("\n=== Instructions ===")
    print("1. Click and drag on the heatmap to select a rectangular region")
    print("2. Release to calculate row-wise averages and plot results")
    print("3. Close the plot window to select another region or exit")
    print("====================\n")
    
    rect_selector = RectangleSelector(ax, onselect, useblit=True, button=[1], 
                                     minspanx=5, minspany=5, spancoords='data', 
                                     use_data_coordinates=True)
    
    plt.show()

if __name__ == "__main__":
    user_input()
