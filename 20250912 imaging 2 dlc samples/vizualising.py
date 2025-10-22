import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl import load_workbook
from matplotlib.widgets import RectangleSelector

def onselect(eclick, erelease):
#    print(eclick,erelease)
   x1, y1 = eclick.xdata, eclick.ydata
   x2, y2 = erelease.xdata, erelease.ydata
   print(f"Selected rectangle coordinates: ({x1}, {y1}) to ({x2}, {y2})")

def visualize_excel_counts(file_path, colourmap, figsize=(12, 10), show_values=False):
    filename = file_path.split("/")[-1]
    title = filename
    wb = load_workbook(file_path)
    sheet_name = wb.sheetnames[0]
    df = pd.read_excel(file_path,sheet_name=sheet_name,header = None)
    df = df.dropna(how='all').dropna(axis=1, how='all')
    df = df.fillna(0)
    df = df.apply(pd.to_numeric, errors='coerce').fillna(0)
    print("Data shape:" , df.shape)
    print("Value range:" , df.values.min() ," to " , df.values.max())
    plt.figure(figsize=figsize)
    ax = sns.heatmap(df,cmap=colourmap,annot=show_values,fmt='.0f', cbar_kws={'label': 'Count'},square=False)
    
    plt.title(title, fontsize=16, pad=20)
    plt.xlabel('Column', fontsize=12)
    plt.ylabel('Row', fontsize=12)
    plt.tight_layout
    print(df[3].values[248])
    return plt,df,ax

    

def check_file(file_path):
    try:
        wb = load_workbook(file_path)
        return True
    except:
        print("no such file, please re-enter")
        return False
def user_input():
    file_correct = False
    while file_correct ==False:
        file_path = input("Enter fileapth and name ( in reference to this directory): ")
        file_correct = check_file(file_path)
    colour_map = input( "choose a colour scheme ('viridis', 'plasma', 'hot', 'coolwarm', 'Blues'): ")
    plt,df,ax = visualize_excel_counts(file_path,colour_map)
    rect_selector = RectangleSelector(ax, onselect, useblit=True, button=[1], minspanx=5, minspany=5, spancoords='data',use_data_coordinates=True)
    
    plt.show()


user_input()